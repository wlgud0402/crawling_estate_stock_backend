from openpyxl import Workbook


def json_estate_to_xlsx(location, estate_json_data):
    estate_json_data = estate_json_data
    wb = Workbook()  # create xlsx file
    ws = wb.active  # create xlsx sheet

    # # 첫번째 줄 헤더
    ws.append([
        "아파트명",
        "건물타입",
        "빌딩타입",
        "동수",
        "세대수",
        "사용승인일",
        "대표이미지",
        "가용매매",
        "가용전세",
        "가용월세",
        "가용단기",
        "총가용",
        "최소면적",
        "최대면적",
        "최소매매가",
        "최대매매가",
        "최소전세가",
        "최대전세가",
    ])
    for i in range(len(estate_json_data)):
        ws.append([estate_json_data[i].get("hscpNm"),
                   estate_json_data[i].get("hscpTypeCd"),
                   estate_json_data[i].get("hscpTypeNm"),
                   estate_json_data[i].get("totDongCnt"),
                   estate_json_data[i].get("totHsehCnt"),
                   estate_json_data[i].get("useAprvYmd"),
                   estate_json_data[i].get("repImgUrl"),
                   estate_json_data[i].get("dealCnt"),
                   estate_json_data[i].get("leaseCnt"),
                   estate_json_data[i].get("rentCnt"),
                   estate_json_data[i].get("strmRentCnt"),
                   estate_json_data[i].get("totalAtclCnt"),
                   estate_json_data[i].get("minSpc"),
                   estate_json_data[i].get("maxSpc"),
                   estate_json_data[i].get("dealPrcMin"),
                   estate_json_data[i].get("dealPrcMax"),
                   estate_json_data[i].get("leasePrcMin"),
                   estate_json_data[i].get("leasePrcMax")])
    wb.save(f'./report/estate/{location}.xlsx')


def json_stock_to_xlsx(kind, page, stock_json_data):
    wb = Workbook()
    ws = wb.active

    ws.append([
        "순위",
        "이름",
        "코드",
        "현재가",
        "시가총액",
        "거래량",
        "전일대비",
        "등락율",
    ])
    for i in range(len(stock_json_data)):
        ws.append([stock_json_data[i].get("rank"),
                   stock_json_data[i].get("name"),
                   stock_json_data[i].get("code"),
                   stock_json_data[i].get("now_price"),
                   stock_json_data[i].get("total_price"),
                   stock_json_data[i].get("trade_amount"),
                   stock_json_data[i].get("compare_yesterday_money"),
                   stock_json_data[i].get("compare_yesterday_proportion")])

    wb.save(f'./report/stock_main/{kind}_{page}.xlsx')


def json_stock_detail_to_xlsx(name, detail_json_data):
    detail_json_data = detail_json_data
    wb = Workbook()
    ws = wb.active

    ws.append([
        "현재가",
        "매도호가",
        "매수호가",
        "등락율",
        "전일대비",
        "전일가",
        "거래량",
        "시가",
        "고가",
        "저가",
        "거래대금(천)",
        "액면가",
        "상한가",
        "하한가",
        "전일상한",
        "전일하한",
        "PER",
        "EPS",
        "high_52",
        "low_52",
        "시가총액",
        "상장주식수",
        "외국인현재",
        "자본금",
    ])
    for i in range(len(detail_json_data)):
        ws.append([detail_json_data[i].get("now_price"),
                   detail_json_data[i].get("sell_price"),
                   detail_json_data[i].get("buy_price"),
                   detail_json_data[i].get("compare_yesterday_proportion"),
                   detail_json_data[i].get("compare_yesterday_money"),
                   detail_json_data[i].get("yesterday_price"),
                   detail_json_data[i].get("trade_amount"),
                   detail_json_data[i].get("start_price"),
                   detail_json_data[i].get("high_price"),
                   detail_json_data[i].get("low_price"),
                   detail_json_data[i].get("income_momey"),
                   detail_json_data[i].get("face_price"),
                   detail_json_data[i].get("high_limit"),
                   detail_json_data[i].get("yesterday_high_limit"),
                   detail_json_data[i].get("low_limit"),
                   detail_json_data[i].get("yesterday_low_limit"),
                   detail_json_data[i].get("PER"),
                   detail_json_data[i].get("EPS"),
                   detail_json_data[i].get("high_52"),
                   detail_json_data[i].get("low_52"),
                   detail_json_data[i].get("total_price"),
                   detail_json_data[i].get("market_sale_count"),
                   detail_json_data[i].get("foreigner"),
                   detail_json_data[i].get("have_momey")])

    wb.save(f'./report/stock_detail/{name}.xlsx')
